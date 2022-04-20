import json
import openpyxl
import string
from pprint import pprint

xl_path = 'active_roster.xlsx'
json_path = 'members.json'
wb = openpyxl.load_workbook(xl_path, read_only=True)
sheet = wb['Active Roster']
db_members = json.loads(open(json_path).read())

xl_member = {
  'first_name': '',
  'last_name': '',
  'id': '',
  'hometown': '',
  'major': '',
  'year': '',
  'club_role': 'active'
}

xl_members = []

for x in range(2,50):
  xl_member = {
  'first_name': sheet.cell(row=x,column=3).value,
  'last_name': sheet.cell(row=x,column=2).value,
  'id': '',
  'hometown': sheet.cell(row=x,column=4).value,
  'major': sheet.cell(row=x,column=6).value,
  'year': sheet.cell(row=x,column=5).value,
  'club_role': 'active'
  }
  xl_member.update({
    'first_name': str(xl_member['first_name']).strip(),
    'hometown': str(xl_member['hometown']).strip(), 
    'major': str(xl_member['major']).strip(),
    'year': str(xl_member['year']).strip(),
  })
  # print(xl_member)
  xl_members.append(xl_member)

# pprint(xl_members)

match = 0
for db_member in db_members:
  # print(db_member)
  for sheet_member in xl_members:
    username = sheet_member['first_name'] + sheet_member['last_name']
    # print(sheet_member['first_name'] + sheet_member['last_name'])
    if username == db_member['user_name']:
      print(f"matched {username} with {db_member['user_name']} id: {db_member['user_id']}")
      sheet_member.update({'id': db_member['user_id']})
      # print(sheet_member)
      match+=1

f_sql = open('update.sql', 'w')
# print(xl_members)

for sheet_member in xl_members:
  if sheet_member['id'] != '': 
    f_sql.write(f"UPDATE member \n\tSET club_role = \'{sheet_member['club_role']}\', home_town = \'{sheet_member['hometown']}\', major = \'{sheet_member['major']}\', year = \'{sheet_member['year']}\' \n ")
    f_sql.write(f"\t\tWHERE user_id = {sheet_member['id']};\n")


f_new_members = open('create_new_mems.sql', 'w')
f_new_members.write('INSERT INTO member(user_name, first_name, last_name, gender, role, password, club_role, home_town, major, year)\n VALUES\n')
for sheet_member in xl_members:
  if sheet_member['id'] == '':
    gender = 'male'
    print(sheet_member['first_name'])
    first_name = sheet_member['first_name']
    last_name = sheet_member['last_name']
    user_name = first_name + last_name
    if first_name == 'Nyla':
      gender = 'female'
    admin = 'user'
    pswd = '123'
    f_new_members.write(f"\t(\'{user_name}\', \'{first_name}\', \'{last_name}\', \'{gender}\', \'{admin}\', \'{pswd}\', \'{sheet_member['club_role']}\', \'{sheet_member['hometown']}\', \'{sheet_member['major']}\', \'{sheet_member['year']}\'),\n")