import json
import openpyxl


def get_roster(roster_sheet):

  roster_list = [] 
  # get first name
  for x in range(2,60):
    first_name = roster_sheet.cell(row=x,column=2).value
    last_name = roster_sheet.cell(row=x,column=3).value
    gender = roster_sheet.cell(row=x,column=4).value
    if gender == 'M':
      gender = 'male'
    elif gender == 'F':
      gender = 'female'
    dict = {
      "first_name": first_name,
      "last_name": last_name,
      "gender": gender
    }
    roster_list.append(dict)
  return(roster_list)

def get_lifts(roster_sheet):
  max_list = []
  for i in range(2,32):
    first_name = roster_sheet.cell(row=i,column=2).value
    last_name = roster_sheet.cell(row=i,column=3).value
    body_weight = roster_sheet.cell(row=i, column=5).value
    squat = roster_sheet.cell(row=i, column=6).value
    bench = roster_sheet.cell(row=i, column=7).value
    deadlift = roster_sheet.cell(row=i, column=8).value
    if deadlift == None or squat == None or bench == None or body_weight == None:
      print(f'no value in row {i}, skipped')
      continue
    dict = {
      "first_name": first_name,
      "last_name": last_name,
      "bench": bench,
      "squat": squat,
      "deadlift": deadlift,
      "body_weight": body_weight
    }
    max_list.append(dict)
  return(max_list)

path = 'data.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
rs = wb['Maxes']

f_members = open('members.json','w')
f_lifts = open('lifts.json', 'w')

f_members.write(json.dumps(get_roster(rs), indent=2))
f_lifts.write(json.dumps(get_lifts(rs),indent=2))

f_members.close()
f_lifts.close()